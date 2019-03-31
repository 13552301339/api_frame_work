from openpyxl import load_workbook
from common.read_config import ReadConfig
from common import file_path


class DoExcel(object):
    def __init__(self, file_name):
        self.file_name = file_name
        # 读取测试用例表单
        self.sheet_dict = eval(ReadConfig().read_config(file_path.config_path, 'CASECONFIG', 'sheet_dict'))

    # 获取测试用例
    def get_data(self):
        wb = load_workbook(self.file_name)

        # 表头
        header = []
        for i in range(1, 8):
            header.append(wb['testcase'].cell(1, i).value)

        test_data = []
        # 定位每个表单
        for sheet_name in self.sheet_dict:
            # print(sheet_name)

            sheet = wb[sheet_name]
            if self.sheet_dict[sheet_name] == 'all':

                # 存储测试用例
                for i in range(2, sheet.max_row+1):
                    # sheet.max_row 最大行
                    # sheet.max_column  最大列
                    subdata = {}
                    for j in range(1, 8):

                        subdata[header[j-1]] = sheet.cell(i, j).value

                    test_data.append(subdata)

            else:
                for i in self.sheet_dict[sheet_name]:
                    subdata = {}
                    for j in range(1, 8):

                        subdata[header[j-1]] = sheet.cell(i+1, j).value
                    test_data.append(subdata)


        # print(test_data)
        return test_data

    # 写回实际结果与测试是否通过
    def weite_back(self, sheet_name, row, ActaulResult, TestResult):

        wb = load_workbook(self.file_name)
        sheet = wb[sheet_name]

        sheet.cell(row, 8).value = ActaulResult
        sheet.cell(row, 9).value = TestResult

        wb.save(self.file_name)


if __name__ == '__main__':

    res = DoExcel(file_path.testcase_path).get_data()
    print(res)

